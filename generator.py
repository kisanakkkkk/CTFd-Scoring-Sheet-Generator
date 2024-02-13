import csv
import openpyxl
import re
import argparse
import os
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

CATEGORIES = []

CHARSET = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

CELL_COLOR = {
    'green': '00FF00',
    'red': 'FF0000',
    'gray': '808080',
    'yellow': 'FFFF00'
}

TITLE = ""

class User:
    def __init__(self, variables):
        for item in variables:
            setattr(self, item, None)
    
    def setData(self, data_list):
        for name, value in zip(self.__dict__, data_list):
            setattr(self, name, value)

class Team:
    def __init__(self, variables):
        for item in variables:
            setattr(self, item, None)
    
    def setData(self, data_list):
        for name, value in zip(self.__dict__, data_list):
            setattr(self, name, value)

class Scoreboard:
    def __init__(self, variables):
        for item in variables:
            setattr(self, item, None)
    
    def setData(self, data_list):
        for name, value in zip(self.__dict__, data_list):
            setattr(self, name, value)

class TeamScoreboard:
    def __init__(self, variables):
        for item in variables:
            setattr(self, item, None)
    
    def setData(self, data_list):
        for name, value in zip(self.__dict__, data_list):
            setattr(self, name, value)

class Challenge:
    def __init__(self, variables):
        for item in variables:
            setattr(self, item, None)
    
    def setData(self, data_list):
        for name, value in zip(self.__dict__, data_list):
            setattr(self, name, value)

class Sheet:
    def __init__(self, place, userid, name, challs, total):
        self.place = place
        self.userid = userid
        self.name = name
        for var_name in challs:
            self.__dict__[var_name] = "X"
        self.total = total

class Accum:
    def __init__(self, place, user, score, categories, final_score, score_change):
        self.place = place
        self.user = user
        self.score = score
        for var_name in categories:
            self.__dict__[var_name] = "X"
        self.final_score = final_score
        self.score_change = score_change

def color_conditional_formatting():
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    gray_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    white_font = Font(color="FFFFFF")

    rule1 = CellIsRule(
        operator="equal",
        formula=['"OK"'],
        stopIfTrue=True,  
        fill=green_fill  

    )

    rule2 = CellIsRule(
        operator="equal",
        formula=['"D"'],  
        stopIfTrue=True,  
        fill=red_fill  
    )

    rule3 = CellIsRule(
        operator="equal",
        formula=['"X"'],  
        stopIfTrue=True,  
        fill=gray_fill,  
        font=white_font
    )

    rule4 = CellIsRule(
        operator="equal",
        formula=['"P"'],  
        stopIfTrue=True,  
        fill=yellow_fill  
    )

    return rule1, rule2, rule3, rule4


def add_data(ishead, sheet, row, dataset, start_col, font=False, alignment=False):
    for col_index, cell_value in enumerate(dataset, start=start_col): # bikin head
        if ishead:
            x = sheet.cell(row=row, column=col_index, value=cell_value)
        else:
            x = sheet.cell(row=row, column=col_index, value=dataset[cell_value])
        if font:
            x.font = font
        if alignment:
            x.alignment = alignment


def add_to_xls(workbook, name, xls_file, classes, heads, color_sheet_range=False):
    name = os.path.basename(name)
    heads_font = Font(name='Lato', size=12, bold=True)
    title_font = Font(name='Lato', size=20, bold=True)
    alg_center = Alignment(horizontal='center', vertical="center")


    if type(heads) is not list:
        heads = vars(heads)
    
    sheet = workbook[name]

    start_row = 1
    if name in CATEGORIES:
        sheet.freeze_panes = sheet['D9']
        workbook.move_sheet(sheet, offset=0)
        #adding legends
        add_data(True, sheet, 2, ["Legends"], 1, font=heads_font)
        add_data(True, sheet, 3, ["OK", "Flag Correct & POC Clear"], 1, font=heads_font)
        add_data(True, sheet, 4, ["D", "Point Deduction"], 1, font=heads_font)
        add_data(True, sheet, 5, ["X", "Not Answered / No Writeup"], 1, font=heads_font)
        add_data(True, sheet, 6, ["P", "Pending / Not Sure"], 1, font=heads_font)
        start_row = 8
    
    if name == "Accumulation":
        workbook.move_sheet(sheet, offset=0)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
        add_data(True, sheet, 1, [TITLE], 1, font=title_font, alignment=alg_center)
        add_data(True, sheet, 2, ["Evaluation Sheet"], 1, font=title_font, alignment=alg_center)
        add_data(True, sheet, 3, ["Link Writeup:", "{{Insert POC Drive Link Here}}"], 1, font=heads_font)
        start_row = 5

    # Populate heads
    add_data(True, sheet, start_row, heads, 1, font=heads_font)
    #fulfill with data, every row, loop every classes
    for row_index, row_data in enumerate(classes, start=start_row + 1):
        add_data(False, sheet, row_index, vars(row_data), 1, alignment=alg_center)

    if color_sheet_range:
        for c in color_conditional_formatting():
            sheet.conditional_formatting.add(color_sheet_range, c)

    for col in sheet.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[get_column_letter(column)].width = adjusted_width


    print(f'{name} imported to {xls_file}')

def add_user(workbook, name, xls_file):
    classes = []
    colnames = []
    with open(name, 'r', encoding="iso-8859-1") as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',', quotechar='|')
        index = 0
        for x in csv_reader:
            newclass = None
            new = []
            for j in x:
                n = j.replace('"', '')
                try:
                    new.append(int(n))
                except ValueError:
                    new.append(n)
            if index == 0:
                colnames = new
            else:
                newclass = User(colnames)
                newclass.setData(new)
                classes.append(newclass)
            index+=1
    pattern = re.compile("[a-z+]*.csv")
    name = re.findall(pattern, os.path.basename(name))[0]
    add_to_xls(workbook, name, xls_file, classes, classes[0])
    return classes

def add_team(workbook, name, xls_file):
    classes = []
    colnames = []
    with open(name, 'r', encoding="iso-8859-1") as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',', quotechar='"')
        index = 0
        for x in csv_reader:
            newclass = None
            new = []
            for j in x:
                n = j
                try:
                    new.append(int(n))
                except ValueError:
                    new.append(n)
            if index == 0:
                colnames = new
            else:
                newclass = Team(colnames)
                newclass.setData(new)
                classes.append(newclass)
            index+=1
    pattern = re.compile("[a-z+]*.csv")
    name = re.findall(pattern, os.path.basename(name))[0]
    add_to_xls(workbook, name, xls_file, classes, classes[0])
    return classes

def add_scoreboard(workbook, name, xls_file):
    classes = []
    colnames = []
    with open(name, 'r', encoding="iso-8859-1") as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',', quotechar='|')
        index = 0
        for x in csv_reader:
            newclass = None
            new = []
            for j in x:
                n = j.replace(';;', '')
                try:
                    new.append(int(n))
                except ValueError:
                    new.append(n)
            if index == 0:
                colnames = new
            else:
                newclass = Scoreboard(colnames)
                newclass.setData(new)
                classes.append(newclass)
            index+=1
    pattern = re.compile("[a-z+]*.csv")
    name = re.findall(pattern, os.path.basename(name))[0]
    add_to_xls(workbook, name, xls_file, classes, classes[0])
    return classes

def add_team_scoreboard(workbook, name, xls_file):
    classes = []
    colnames = []
    with open(name, 'r', encoding="iso-8859-1") as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',', quotechar='"')
        index = 0
        for x in csv_reader:
            newclass = None
            new = []
            for j in x:
                n = j
                try:
                    new.append(int(n))
                except ValueError:
                    new.append(n)
            if index == 0:
                colnames = new
            else:
                newclass = TeamScoreboard(colnames)
                newclass.setData(new)
                classes.append(newclass)
            index+=1
    pattern = re.compile("[a-z+]*.csv")
    name = re.findall(pattern, os.path.basename(name))[0]
    add_to_xls(workbook, name, xls_file, classes, classes[0])
    return classes

def add_chall(workbook, name, xls_file, files):
    global CATEGORIES
    classes = []
    colnames = []
    with open(name, 'r', encoding="iso-8859-1") as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',', quotechar='"')
        index = 0
        for x in csv_reader:
            newclass = None
            new = []
            for j in x:
                n = j.replace(';;', '')
                try:
                    new.append(int(n))
                except ValueError:
                    new.append(n)
            if index == 0:
                colnames = new
            else:
                newclass = Challenge(colnames)
                newclass.setData(new)
                classes.append(newclass)
            index+=1
    
    chall_category = set()
    for cl in classes:
        chall_category.add(cl.category)
    
    CATEGORIES = sorted(chall_category)
    preparing_sheets(workbook, files)
    pattern = re.compile("[a-z+]*.csv")
    name = re.findall(pattern, os.path.basename(name))[0]
    add_to_xls(workbook, name, xls_file, classes, classes[0])
    return classes

def sanitize(files_path):
    folder = "sanitized"
    if not os.path.exists(folder):
        os.makedirs(folder)
    new_file = []
    try:
        for csv_file_path in files_path:            
            f = open(csv_file_path, 'rb')
            fu = f.read()
            f.close()
            new_name = folder + '/san_' + csv_file_path.strip('.\\')
            f = open(new_name, 'wb')
            if b'\0' in fu:
                fu = fu.replace(b'\0', b'')
            f.write(fu)
            f.close()
            new_file.append(new_name)
        return new_file
    except Exception as e:
        print('Error in sanitizing files', e)
        exit()

def create_accumulation_sheet(workbook, xls_file, scoreboard_classes, chall_classes, isTeam):
    accumulation_classes = []
    index = 0
    sc_class = []
    for sc in scoreboard_classes:
        if sc.place != '':
            sc_class.append(sc)

    for c in sc_class:
        place = c.place
        score = c.score
        user = None
        if isTeam:
            user = c.team
        else:
            user = c.user
        formula_final = f"=SUM(D{6+index}:{CHARSET[3 + len(CATEGORIES) - 1]}{6+index})"
        formula_scorechange = f"=NOT(C{6+index}={CHARSET[3 + len(CATEGORIES) - 1 + 1]}{6+index})"
        accumulation_classes.append(Accum(place, user, score, CATEGORIES, formula_final, formula_scorechange))
        index+=1
    heads = ["Place","User","CTFd Score"] + CATEGORIES + ["Final Score", "Score Change?"]

    nums = {}
    for x in chall_classes:
        if x.category in nums:
            nums[x.category] += 1
        else:
            nums[x.category] = 1

    index = 0
    for ac in accumulation_classes:
        for ca in CATEGORIES:
            formula_category = f"=VLOOKUP($B{6+index};'{ca}'!$C$9:${CHARSET[nums[ca] + 4 - 1]}${9 + (len(accumulation_classes) - 1)};{nums[ca]+2};FALSE)"
            ac.__dict__[ca] = formula_category
        index += 1
    add_to_xls(workbook, "Accumulation", xls_file, accumulation_classes, heads)

def create_category_sheet(workbook, category_name, xls_file, scoreboard_classes, entrantclasses, chall_classes, isTeam):
    chall_names = []
    for x in chall_classes:
        if x.category == category_name:
            chall_names.append(x.name)
    sheet_classes = []
    index = 0
    
    sc_class = []
    for c in scoreboard_classes:
        if c.place != '':
            sc_class.append(c)

    for sc in sc_class:
        for us in entrantclasses:
            user = None
            if isTeam:
                user = sc.team
            else:
                user = sc.user
            place = sc.place
            if user == us.name:
                id = us.id
                formula_total = f'=SUMPRODUCT($D${str(9 + len(sc_class))}:${chr(ord("D") + len(chall_names) - 1)}${str(9 + len(sc_class))}, MMULT({{1,0.5,0,0}},--(D{str(9+index)}:{chr(ord("D") + len(chall_names) - 1)}{str(9+index)}={{"OK";"D";"X";"P"}})))'
                sheet_classes.append(Sheet(place, id, user, chall_names, formula_total))
                index+=1
    
    scores = Sheet(None, None, None, chall_names, None)
    for x in chall_classes:
        if x.category == category_name:
            scores.__dict__[x.name] = x.value

    sheet_classes = sheet_classes + [scores]
    
    heads = ['Place', 'ID', 'User'] + chall_names + ['Total']
    ranges = f"D9:{CHARSET[len(chall_names) + 3]}{8 + len(sc_class)}"
    add_to_xls(workbook, category_name, xls_file, sheet_classes, heads, ranges)

def preparing_sheets(workbook, files):
    pattern = re.compile("[a-z+]*.csv")
    names = [re.findall(pattern, os.path.basename(f))[0] for f in files]
    sheets = ["Accumulation"] + CATEGORIES + names
    for s in sheets:
        sheet = workbook.create_sheet(f'{s}')

    #delete initial sheet
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])
    print('Preparation Finished')

def generate(args):
    global TITLE

    entrantcsv = args.data
    scorecsv = args.score
    challcsv = args.chall
    isTeam = args.team

    TITLE = os.path.basename(entrantcsv).split('-')[0]
    xls_file = TITLE + "-scoringsheet.xlsx"
    if args.output:
        xls_file = args.output

    workbook = openpyxl.Workbook()
    #creating sanitized copy
    entrantcsv, scorecsv, challcsv = sanitize([entrantcsv, scorecsv, challcsv])

    chall_classes = add_chall(workbook, challcsv, xls_file, [entrantcsv, scorecsv, challcsv])
    if isTeam:
        entrantclasses = add_team(workbook, entrantcsv, xls_file)
        scoreboard_classes = add_team_scoreboard(workbook, scorecsv, xls_file)
    else:
        entrantclasses = add_user(workbook, entrantcsv, xls_file)
        scoreboard_classes = add_scoreboard(workbook, scorecsv, xls_file)
    create_accumulation_sheet(workbook, xls_file, scoreboard_classes, chall_classes, isTeam)
    for category in CATEGORIES:
        create_category_sheet(workbook, category, xls_file, scoreboard_classes, entrantclasses, chall_classes, isTeam)
    
    workbook.save(xls_file)

def main():
    parser = argparse.ArgumentParser()

    parser.add_argument('-d', '--data', type=str, required=True, help='specify user/team data CSV file')
    parser.add_argument('-s', '--score', type=str, required=True, help='specify scoreboard CSV file')
    parser.add_argument('-c', '--chall', type=str, required=True, help='specify challenges CSV file')
    parser.add_argument('-o', '--output', type=str, help='set output file (default: output.xlsx)')
    parser.add_argument('-t', '--team', action='store_true', help='indicates that the CTF is team-based (default: individual-based)')

    args = parser.parse_args()

    if not args.data and args.score and args.chall:
        parser.error('Please specify corresponding files')
    generate(args)

if __name__ == '__main__':
    main()